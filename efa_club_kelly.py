"""
EFA Kelly formula helpers — sourced from `EFA - Kelly Formula.xlsx`.

Workbook (Sheet1):
  B3  classic binary Kelly:  f* = (p * b - q) / b
  C14 Kelly Adjustment (half-Kelly default = 0.5)
  C19 workbook equation:     adj * ((p * (b + 1) - q) / bankroll)
  B21 continuous Kelly:      f* = (μ - r) / σ²
"""
from __future__ import annotations

from pathlib import Path

DEFAULT_WIN_PROBS = (0.55, 0.65, 0.70)
DEFAULT_NET_ODDS = (1.0, 1.5, 3.0)
DEFAULT_BANKROLLS = (1.0, 2.0, 3.0, 5.0)
DEFAULT_KELLY_ADJUSTMENT = 0.5

KELLY_WORKBOOK_NAMES = ("EFA - Kelly Formula.xlsx", "EFA_Kelly_Formula.xlsx", "kelly.xlsx")


def find_kelly_workbook(base_dir=None):
    roots = []
    if base_dir:
        roots.append(Path(base_dir))
    roots.append(Path(__file__).resolve().parent)
    home = Path.home()
    roots.extend(
        [
            home / "Documents" / "EFA_Club",
            home / "Documents" / "documents" / "EFA_Club",
        ]
    )
    seen = set()
    for root in roots:
        try:
            resolved = root.resolve()
        except OSError:
            continue
        if resolved in seen:
            continue
        seen.add(resolved)
        for name in KELLY_WORKBOOK_NAMES:
            path = resolved / name
            if path.is_file():
                return path
    return None


def load_workbook_formulas(path=None):
    """Read formula labels from the club Kelly workbook. Falls back to hardcoded text."""
    fallback = {
        "path": None,
        "binary_formula": "f* = (p * b - q) / b",
        "binary_terms": {
            "f*": "fraction of bankroll to invest (Kelly Fraction)",
            "p": "Probability of winning",
            "q": "1 - p = probability of losing",
            "b": "net odds received per unit wagered (+100 odds means b = 1)",
        },
        "kelly_adjustment": DEFAULT_KELLY_ADJUSTMENT,
        "workbook_equation": "adj * ((p * (b + 1) - q) / bankroll)",
        "continuous_formula": "f* = (μ - r) / σ²",
        "continuous_terms": {
            "μ": "Expected excess return of the asset above the risk-free rate",
            "r": "risk-free rate",
            "σ²": "Variance (risk) of the asset's returns",
        },
        "notes": [
            "Half-Kelly uses the workbook Kelly Adjustment of 0.5.",
            "Workbook cell C19 divides by bankroll rather than by net odds.",
            "Classic Kelly (B3) is used for the club position-size matrix.",
        ],
    }
    path = Path(path) if path else find_kelly_workbook()
    if path is None:
        return fallback
    try:
        import openpyxl
    except ImportError:
        fallback["path"] = str(path)
        fallback["notes"] = list(fallback["notes"]) + [
            f"Workbook found at {path.name} but openpyxl is not installed; using built-in formulas."
        ]
        return fallback
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
        ws = wb[wb.sheetnames[0]]
        values = {}
        for row in ws.iter_rows(min_row=1, max_row=35, max_col=4):
            for cell in row:
                if cell.value is not None:
                    values[cell.coordinate] = cell.value
        wb.close()
    except Exception as e:
        fallback["path"] = str(path)
        fallback["notes"] = list(fallback["notes"]) + [f"Could not parse workbook: {e}"]
        return fallback

    adj = values.get("C14")
    try:
        adj = float(adj) if adj is not None else DEFAULT_KELLY_ADJUSTMENT
    except (TypeError, ValueError):
        adj = DEFAULT_KELLY_ADJUSTMENT

    return {
        "path": str(path),
        "binary_formula": str(values.get("B3") or fallback["binary_formula"]),
        "binary_terms": {
            "f*": str(values.get("C4") or fallback["binary_terms"]["f*"]),
            "p": str(values.get("C5") or fallback["binary_terms"]["p"]),
            "q": str(values.get("C6") or fallback["binary_terms"]["q"]),
            "b": str(values.get("C7") or fallback["binary_terms"]["b"]),
        },
        "kelly_adjustment": adj,
        "workbook_equation": str(values.get("C19") or fallback["workbook_equation"]),
        "continuous_formula": str(values.get("B21") or fallback["continuous_formula"]),
        "continuous_terms": {
            "μ": str(values.get("C23") or fallback["continuous_terms"]["μ"]),
            "r": str(values.get("C24") or fallback["continuous_terms"]["r"]),
            "σ²": str(values.get("C25") or fallback["continuous_terms"]["σ²"]),
        },
        "notes": [
            str(values.get("D14") or "Adjust for 1/4, 1/2, 3/4, and full Kelly"),
            str(values.get("D16") or "Net odds: 1 = +100, 1.5 = +150, 2 = +200"),
            str(values.get("D17") or "Bankroll size scales the suggested stake."),
            "Classic Kelly (B3) is used for the club half-Kelly matrix.",
            "Workbook C19 is shown alongside for comparison.",
        ],
    }


def loss_probability(win_probability):
    return 1.0 - float(win_probability)


def classic_kelly_fraction(win_probability, net_odds):
    """B3: f* = (p * b - q) / b."""
    p = float(win_probability)
    b = float(net_odds)
    if b == 0:
        return 0.0
    q = loss_probability(p)
    return (p * b - q) / b


def half_kelly_fraction(win_probability, net_odds, adjustment=DEFAULT_KELLY_ADJUSTMENT):
    return float(adjustment) * classic_kelly_fraction(win_probability, net_odds)


def suggested_stake(win_probability, net_odds, bankroll, adjustment=DEFAULT_KELLY_ADJUSTMENT):
    """Half-Kelly fraction of bankroll (classic B3 × adjustment × bankroll)."""
    return half_kelly_fraction(win_probability, net_odds, adjustment) * float(bankroll)


def workbook_c19(win_probability, net_odds, bankroll, adjustment=DEFAULT_KELLY_ADJUSTMENT):
    """
    Excel C19: adj * ((p * (b + 1) - (1 - p)) / bankroll)

    This is more aggressive than classic Kelly because (p*(b+1) - q)
    equals (p*b - q) + p, not (p*(b+1) - 1).
    """
    p = float(win_probability)
    b = float(net_odds)
    br = float(bankroll)
    if br == 0:
        return 0.0
    q = loss_probability(p)
    return float(adjustment) * ((p * (b + 1) - q) / br)


def continuous_kelly(expected_return, risk_free_rate, variance):
    """B21 / C31: f* = (μ - r) / σ²."""
    var = float(variance)
    if var == 0:
        return 0.0
    return (float(expected_return) - float(risk_free_rate)) / var


def nearest_grid_value(value, grid):
    if value is None:
        return None
    try:
        target = float(value)
    except (TypeError, ValueError):
        return None
    return min(grid, key=lambda item: abs(float(item) - target))


def build_half_kelly_matrix(
    win_probs=DEFAULT_WIN_PROBS,
    net_odds=DEFAULT_NET_ODDS,
    bankrolls=DEFAULT_BANKROLLS,
    adjustment=DEFAULT_KELLY_ADJUSTMENT,
):
    """Full 3-way matrix used on the assessment tab and printable report."""
    fraction_grid = []
    for p in win_probs:
        row = {"win_probability": float(p)}
        for b in net_odds:
            f_full = classic_kelly_fraction(p, b)
            f_half = half_kelly_fraction(p, b, adjustment)
            row[f"odds_{b:g}"] = {
                "net_odds": float(b),
                "full_kelly": f_full,
                "half_kelly": f_half,
            }
        fraction_grid.append(row)

    stake_rows = []
    for p in win_probs:
        for b in net_odds:
            f_half = half_kelly_fraction(p, b, adjustment)
            for br in bankrolls:
                stake_rows.append(
                    {
                        "win_probability": float(p),
                        "net_odds": float(b),
                        "loss_probability": loss_probability(p),
                        "full_kelly": classic_kelly_fraction(p, b),
                        "half_kelly": f_half,
                        "bankroll": float(br),
                        "suggested_stake": f_half * float(br),
                        "workbook_c19": workbook_c19(p, b, br, adjustment),
                    }
                )

    return {
        "adjustment": float(adjustment),
        "win_probs": [float(p) for p in win_probs],
        "net_odds": [float(b) for b in net_odds],
        "bankrolls": [float(br) for br in bankrolls],
        "fraction_grid": fraction_grid,
        "stake_rows": stake_rows,
        "formula": {
            "classic": "f* = (p * b - q) / b",
            "half": "f½ = 0.5 × (p * b - q) / b",
            "stake": "suggested units = f½ × bankroll",
            "workbook_c19": "0.5 × ((p * (b + 1) - q) / bankroll)",
        },
    }


def format_pct(value, digits=1):
    try:
        return f"{float(value) * 100:.{digits}f}%"
    except (TypeError, ValueError):
        return "n/a"


def format_units(value, digits=3):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "n/a"
    if number <= 0:
        return "0 (no edge)"
    return f"{number:.{digits}f}"


def matrix_to_fraction_table(matrix):
    rows = []
    for item in matrix.get("fraction_grid") or []:
        row = {"Win probability": f"{item['win_probability'] * 100:.0f}%"}
        for cell in item.values():
            if not isinstance(cell, dict):
                continue
            odds = cell["net_odds"]
            row[f"Half-Kelly @ {odds:g}"] = format_pct(cell["half_kelly"])
            row[f"Full Kelly @ {odds:g}"] = format_pct(cell["full_kelly"])
        rows.append(row)
    return rows


def matrix_to_stake_table(matrix):
    rows = []
    for item in matrix.get("stake_rows") or []:
        rows.append(
            {
                "Win p": f"{item['win_probability'] * 100:.0f}%",
                "Net odds b": f"{item['net_odds']:g}",
                "Half-Kelly f½": format_pct(item["half_kelly"]),
                "Bankroll": f"{item['bankroll']:g}",
                "Suggested units (f½ × BR)": format_units(item["suggested_stake"]),
                "Workbook C19": format_units(item["workbook_c19"]),
            }
        )
    return rows


def highlight_cell(matrix, win_probability=None, net_odds=None):
    if win_probability is None or net_odds is None:
        return None
    p = nearest_grid_value(win_probability, matrix["win_probs"])
    b = nearest_grid_value(net_odds, matrix["net_odds"])
    if p is None or b is None:
        return None
    return {
        "win_probability": p,
        "net_odds": b,
        "half_kelly": half_kelly_fraction(p, b, matrix["adjustment"]),
        "full_kelly": classic_kelly_fraction(p, b),
    }
